#! python3
# update_produce.py - корректирует цены в эл. таблице данных об объемах продаж

import openpyxl
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font

wb = openpyxl.load_workbook('produceSales.xlsx')
sheet = wb.get_sheet_by_name('Sheet')

# типы продукции и их обновленные цены
PRICE_UPDATES = {'Garlic': 3.07,
                 'Celery': 1.19,
                 'Lemon': 1.27}

# стиль измененных ячеек
bold24Font = Font(size=24, bold=True, name='Tahoma')

for row_num in range(2, sheet.max_row): # 1я строка игнорируется
    produce_name = sheet.cell(row=row_num, column=1).value 
    if produce_name in PRICE_UPDATES:
        sheet.cell(row=row_num, column=2).value = PRICE_UPDATES[produce_name]
        sheet.cell(row=row_num, column=2).font = bold24Font

sheet['E1'] = '=SUM(B2:B10000)'
sheet['E2'] = '=SUM(C2:C10000)'
sheet['E3'] = '=SUM(D2:D10000)'

# настройка высоты строк и ширины столбцов

sheet.row_dimensions[1].height = 25
sheet.column_dimensions['E'].width = 35

# слияние и отмена слияния ячеек
sheet.merge_cells('F1:H6')
sheet.unmerge_cells('F1:H6')

wb.save('updatedProduceSales.xlsx')

wb_data_only = openpyxl.load_workbook('updatedProduceSales.xlsx')
sheet = wb_data_only.active
print(sheet['E1'].value)
wb_data_only = openpyxl.load_workbook('updatedProduceSales.xlsx', data_only=True)
sheet = wb_data_only.active
print(sheet['E1'].value)


