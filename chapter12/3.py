#выполнение преобразований между буквенными и цифровыми обозначениями столбцов
import openpyxl

from openpyxl.utils import get_column_letter, column_index_from_string

print(get_column_letter(1))
print(get_column_letter(2))
print(get_column_letter(900))

wb = openpyxl.load_workbook('example.xlsx')
sheet = wb.get_sheet_by_name('Лист1')
print(get_column_letter(sheet.max_column))
print(column_index_from_string('A'))
print(column_index_from_string('AA'))

#получение строк и столбцов рабочих листов
print(tuple(sheet['A1': 'C3']))
for row_of_cell_obj in sheet['A1': 'C3']:
    for cell_obj in row_of_cell_obj:
        print(cell_obj.coordinate, cell_obj.value)
    print('---end---')

print('----=========----')

