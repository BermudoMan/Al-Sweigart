import openpyxl
wb = openpyxl.load_workbook('example.xlsx')
sheet = wb.active
sheet.title = 'Spam Spam Spam'
# создание, добавление новых листов
wb.create_sheet()
wb.create_sheet(index=0, title='1')
wb.create_sheet(index=2, title='medium')
print(wb.get_sheet_names())
wb.remove_sheet(wb.get_sheet_by_name('medium'))
print(wb.get_sheet_names())
wb.save('example_copy.xlsx')

# запись значений в ячейки
wb = openpyxl.load_workbook('example.xlsx')
sheet = wb.get_sheet_by_name('Лист1')
sheet['D1'] = 'HELLO WORLD!'
print(sheet['D1'].value)
wb.save('example_copy2.xlsx')