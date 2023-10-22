import openpyxl
import os

wb = openpyxl.load_workbook('example.xlsx')
sheet = wb.get_sheet_by_name('Лист1')
print(sheet['A1'])
print(sheet['A1'].value)

c = sheet['B1']
print(c.value)

print('Cтрока' + str(c.row) + ', Столбец ' + str(c.column) + ': ' + c.value)

print('Ячейка ' + c.coordinate + ' : ' + c.value)

print(sheet['C1'].value)
#альтернативная индексация строк-столбцов
print(sheet.cell(row=1, column=2))
print(sheet.cell(row=1, column=2).value)

for i in range(1, 8, 2):
    print(i, sheet.cell(row=i, column=2).value)


print('---------')
print(sheet.max_row)
print(sheet.max_column)